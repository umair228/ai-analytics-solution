"""Phase 5 tests: deterministic chart selection, AnswerEnvelope assembly,
server-side exporters, and the ask-live / export endpoints (agent mocked)."""
import io
from unittest import mock

from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase
from rest_framework.test import APIClient

from ai.agent import AgentRun
from ai.envelope import (AnswerEnvelope, envelope_from_agent_run,
                         envelope_from_result, suggest_chart)
from ai.exporters import ExportUnavailable, export, to_csv, to_html, to_xlsx


class SuggestChartTests(SimpleTestCase):
    def test_scalar_kpi(self):
        self.assertEqual(suggest_chart(["total"], [[42]]), {"type": "kpi", "value_field": "total"})

    def test_temporal_line(self):
        c = suggest_chart(["month", "n"], [["2026-01", 5], ["2026-02", 8]])
        self.assertEqual(c["type"], "line")

    def test_few_categories_pie(self):
        c = suggest_chart(["product", "n"], [["A", 1], ["B", 2], ["C", 3]])
        self.assertEqual(c["type"], "pie")

    def test_many_categories_bar(self):
        rows = [[f"P{i}", i] for i in range(12)]
        self.assertEqual(suggest_chart(["product", "n"], rows)["type"], "bar")

    def test_non_numeric_measure_is_table(self):
        self.assertIsNone(suggest_chart(["product", "grade"], [["A", "hi"], ["B", "lo"]]))

    def test_wide_is_table(self):
        self.assertIsNone(suggest_chart(["a", "b", "c"], [[1, 2, 3]]))

    def test_empty_is_none(self):
        self.assertIsNone(suggest_chart([], []))
        self.assertIsNone(suggest_chart(["x"], []))


class EnvelopeTests(SimpleTestCase):
    def test_from_result_builds_chart_and_provenance(self):
        result = {
            "understood": True, "columns": ["product", "off_spec_pct"],
            "rows": [["PROPANE", 36.0], ["LPG", 12.4]], "row_count": 2, "truncated": False,
            "sql": "SELECT ...", "path": "generated", "freshness": "live",
            "datasource": {"id": 3, "name": "EGPC"}}
        env = envelope_from_result(result, narrative="PROPANE worst at 36%.")
        self.assertTrue(env.understood)
        self.assertEqual(env.chart["type"], "pie")
        self.assertEqual(env.provenance["sql"], "SELECT ...")
        self.assertEqual(env.provenance["path"], "generated")
        self.assertIn("generated_at", env.provenance)
        self.assertEqual(env.result["row_count"], 2)

    def test_from_failed_result_surfaces_error_and_suggestions(self):
        env = envelope_from_result(
            {"understood": False, "error": "no valid query", "available_tables": ["SAMPLE"]})
        self.assertFalse(env.understood)
        self.assertEqual(env.error, "no valid query")
        self.assertEqual(env.suggestions, ["SAMPLE"])
        self.assertIsNone(env.chart)

    def test_from_agent_run_extracts_ask_database_table(self):
        run = AgentRun(
            answer="PROPANE has the worst rate.",
            trace=[
                {"tool": "list_datasets", "result": {"datasets": []}},
                {"tool": "ask_database", "result": {
                    "columns": ["product", "pct"], "rows": [["PROPANE", 36.0], ["LPG", 12.0]],
                    "row_count": 2, "truncated": False, "sql": "SELECT ...", "path": "generated"}},
            ],
            steps=2, tool_calls=2, stopped_reason="answered")
        env = envelope_from_agent_run(run, model="qwen2.5:14b", user_id=7)
        self.assertEqual(env.narrative, "PROPANE has the worst rate.")
        self.assertEqual(env.result["columns"], ["product", "pct"])
        self.assertEqual(env.chart["type"], "pie")
        self.assertEqual(env.provenance["sql"], "SELECT ...")
        self.assertEqual(env.provenance["model"], "qwen2.5:14b")
        self.assertEqual(env.provenance["user_id"], 7)

    def test_from_agent_run_no_table_is_narrative_only(self):
        run = AgentRun(answer="I could not find that.", trace=[], steps=1,
                       tool_calls=0, stopped_reason="answered")
        env = envelope_from_agent_run(run, model="m")
        self.assertEqual(env.result["rows"], [])
        self.assertIsNone(env.chart)

    def test_from_agent_run_propagates_failed_ask_database(self):
        run = AgentRun(
            answer="I can't answer that from the available tables.",
            trace=[{"tool": "ask_database", "result": {
                "understood": False, "error": "Could not produce a valid query.",
                "available_tables": ["SAMPLE", "RESULT"]}}],
            steps=2, tool_calls=1, stopped_reason="answered")
        env = envelope_from_agent_run(run, model="m")
        self.assertFalse(env.understood)
        self.assertEqual(env.error, "Could not produce a valid query.")
        self.assertEqual(env.suggestions, ["SAMPLE", "RESULT"])


def _envelope_dict():
    return AnswerEnvelope(
        narrative="PROPANE worst at 36%.",
        result={"columns": ["product", "pct"], "rows": [["PROPANE", 36.0], ["LPG", 12.4]],
                "row_count": 2, "truncated": False},
        chart={"type": "bar", "category_field": "product", "value_field": "pct"},
        provenance={"sql": "SELECT product, pct FROM x", "path": "generated",
                    "freshness": "live", "model": "qwen2.5:14b"}).as_dict()


class ExporterTests(SimpleTestCase):
    def test_csv_has_header_rows_and_provenance(self):
        out = to_csv(_envelope_dict()).decode("utf-8")
        self.assertIn("product,pct", out)
        self.assertIn("PROPANE,36.0", out)
        self.assertIn("# sql", out)

    def test_html_has_narrative_and_provenance(self):
        out = to_html(_envelope_dict())
        self.assertIn("PROPANE worst at 36%.", out)
        self.assertIn("Provenance", out)
        self.assertIn("SELECT product, pct FROM x", out)

    def test_xlsx_is_valid_and_has_data(self):
        data = to_xlsx(_envelope_dict())
        wb = load_workbook(io.BytesIO(data))
        self.assertIn("Provenance", wb.sheetnames)
        flat = [c for row in wb["Answer"].iter_rows(values_only=True) for c in row]
        self.assertIn("PROPANE", flat)

    def test_xlsx_tolerates_nonprimitive_cells(self):
        # a client-POSTed envelope could carry a nested cell — must not crash openpyxl
        env = _envelope_dict()
        env["result"]["rows"] = [["X", {"nested": 1}], ["Y", [1, 2]]]
        wb = load_workbook(io.BytesIO(to_xlsx(env)))  # should not raise
        self.assertEqual(wb.active.title, "Answer")


class SerializationTests(SimpleTestCase):
    def test_non_finite_floats_become_none(self):
        from core.serialization import coerce_value, jsonable_rows
        self.assertIsNone(coerce_value(float("nan")))
        self.assertIsNone(coerce_value(float("inf")))
        self.assertEqual(coerce_value(3.5), 3.5)
        # and a NaN-bearing row renders through DRF's JSON renderer without error
        from rest_framework.renderers import JSONRenderer
        rows = jsonable_rows([[1, float("nan")]])
        JSONRenderer().render({"rows": rows})  # must not raise ValueError

    def test_export_dispatch_and_unknown_format(self):
        data, ctype, fname = export(_envelope_dict(), "csv")
        self.assertTrue(fname.endswith(".csv"))
        self.assertEqual(ctype, "text/csv")
        with self.assertRaises(ExportUnavailable):
            export(_envelope_dict(), "docx")


class EndpointTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username="ep_t", password="x")
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def test_ask_live_returns_envelope(self):
        run = AgentRun(
            answer="PROPANE worst.",
            trace=[{"tool": "ask_database", "result": {
                "columns": ["product", "pct"], "rows": [["PROPANE", 36.0]],
                "row_count": 1, "truncated": False, "sql": "SELECT ...", "path": "generated"}}],
            steps=1, tool_calls=1, stopped_reason="answered")
        with mock.patch("ai.views.is_configured", return_value=True), \
                mock.patch("ai.views.active_model", return_value="qwen2.5:14b"), \
                mock.patch("ai.views.run_agent", return_value=run):
            resp = self.client.post("/api/ai/ask-live/", {"message": "worst product?"}, format="json")
        self.assertEqual(resp.status_code, 200)
        env = resp.json()["envelope"]
        self.assertEqual(env["narrative"], "PROPANE worst.")
        self.assertEqual(env["provenance"]["sql"], "SELECT ...")
        self.assertEqual(env["result"]["row_count"], 1)

    def test_ask_live_requires_message(self):
        with mock.patch("ai.views.is_configured", return_value=True):
            resp = self.client.post("/api/ai/ask-live/", {}, format="json")
        self.assertEqual(resp.status_code, 400)

    def test_export_endpoint_returns_csv(self):
        resp = self.client.post("/api/ai/export/",
                                {"envelope": _envelope_dict(), "format": "csv"}, format="json")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/csv")
        self.assertIn(b"PROPANE", resp.content)

    def test_export_pdf_without_weasyprint_is_clean_400(self):
        # WeasyPrint isn't installed in this env → a clean 400, not a 500.
        try:
            import weasyprint  # noqa: F401
            self.skipTest("WeasyPrint installed; the unavailable-path test does not apply.")
        except ModuleNotFoundError:
            pass
        resp = self.client.post("/api/ai/export/",
                                {"envelope": _envelope_dict(), "format": "pdf"}, format="json")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("WeasyPrint", resp.json()["detail"])
