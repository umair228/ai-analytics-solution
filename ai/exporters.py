"""Server-side report export from an AnswerEnvelope (DIS_TalkToData_Plan.md §6).

Every export carries the provenance footer (exact SQL, path, freshness) so
a regulated-lab reader can reproduce the number. CSV / XLSX / HTML are always
available; PDF needs WeasyPrint (falls back to a clear, catchable error so callers
can offer HTML instead).
"""
from __future__ import annotations

import csv
import html
import io
import json


class ExportUnavailable(Exception):
    """A requested export format needs an optional dependency that isn't installed."""


def _xlsx_cell(value):
    """openpyxl only accepts primitives; stringify anything else (defensive — a
    client-POSTed envelope could carry nested cells)."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, default=str)


def _rows(envelope):
    res = envelope.get("result") or {}
    return res.get("columns") or [], res.get("rows") or []


def _provenance_lines(envelope):
    p = envelope.get("provenance") or {}
    order = ["path", "freshness", "data_as_of", "datasource", "agent_steps",
             "sql", "generated_at", "user_id"]
    # "model" is deliberately excluded: reports are end-user facing and must not
    # name the underlying LLM.
    hidden = {"model"}
    out = []
    for k in order:
        if p.get(k) not in (None, ""):
            out.append((k, p[k]))
    for k, v in p.items():
        if k not in hidden and k not in dict(out) and v not in (None, ""):
            out.append((k, v))
    return out


def to_csv(envelope) -> bytes:
    columns, rows = _rows(envelope)
    buf = io.StringIO()
    w = csv.writer(buf)
    if columns:
        w.writerow(columns)
    for r in rows:
        w.writerow(list(r))
    # provenance as trailing comment rows
    buf.write("\n")
    for k, v in _provenance_lines(envelope):
        w.writerow([f"# {k}", v])
    return buf.getvalue().encode("utf-8")


def to_html(envelope) -> str:
    columns, rows = _rows(envelope)
    narrative = html.escape(envelope.get("narrative") or "")
    head = "".join(f"<th>{html.escape(str(c))}</th>" for c in columns)
    body = "".join(
        "<tr>" + "".join(f"<td>{html.escape('' if c is None else str(c))}</td>" for c in r) + "</tr>"
        for r in rows)
    prov = "".join(
        f"<tr><td class='k'>{html.escape(str(k))}</td><td>{html.escape(str(v))}</td></tr>"
        for k, v in _provenance_lines(envelope))
    return f"""<!doctype html><html><head><meta charset="utf-8"><style>
      body{{font-family:Arial,Helvetica,sans-serif;color:#16202b;margin:24px;}}
      h1{{font-size:16pt;color:#0d4d7a;}} p.narr{{font-size:11pt;}}
      table{{border-collapse:collapse;width:100%;margin:10px 0;font-size:9.5pt;}}
      th,td{{border:1px solid #d7dee6;padding:5px 8px;text-align:left;}}
      th{{background:#0d4d7a;color:#fff;}}
      .prov{{margin-top:18px;font-size:8.5pt;color:#5b6b7b;}}
      .prov td.k{{font-weight:bold;width:120px;}} code{{word-break:break-all;}}
    </style></head><body>
      <h1>DIS — Answer Report</h1>
      <p class="narr">{narrative}</p>
      <table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>
      <div class="prov"><b>Provenance</b><table>{prov}</table></div>
    </body></html>"""


def to_xlsx(envelope) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ExportUnavailable("XLSX export requires openpyxl.") from exc

    columns, rows = _rows(envelope)
    wb = Workbook()
    ws = wb.active
    ws.title = "Answer"
    if envelope.get("narrative"):
        ws.append([envelope["narrative"]])
        ws.append([])
    header_row = ws.max_row + 1
    if columns:
        ws.append([str(c) for c in columns])
    for r in rows:
        ws.append([_xlsx_cell(c) for c in r])

    # A native chart when the spec + data support it (2-col numeric).
    chart_spec = envelope.get("chart") or {}
    ctype = chart_spec.get("type")
    if ctype in ("bar", "line", "pie") and len(columns) == 2 and rows:
        data_last = header_row + len(rows)
        cls = {"bar": BarChart, "line": LineChart, "pie": PieChart}[ctype]
        chart = cls()
        chart.title = chart_spec.get("value_field") or "value"
        data = Reference(ws, min_col=2, min_row=header_row, max_row=data_last)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=data_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"E{header_row}")

    prov = wb.create_sheet("Provenance")
    for k, v in _provenance_lines(envelope):
        prov.append([k, str(v)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(envelope) -> bytes:
    try:
        from weasyprint import HTML  # type: ignore
    except ModuleNotFoundError as exc:
        raise ExportUnavailable(
            "PDF export requires WeasyPrint (pip install weasyprint). "
            "Use HTML export and print to PDF, or install WeasyPrint on the server."
        ) from exc
    return HTML(string=to_html(envelope)).write_pdf()


_EXPORTERS = {
    "csv": (to_csv, "text/csv"),
    "xlsx": (to_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "html": (lambda e: to_html(e).encode("utf-8"), "text/html"),
    "pdf": (to_pdf, "application/pdf"),
}


def export(envelope, fmt):
    """Return (bytes, content_type, filename) for ``fmt`` in {csv,xlsx,html,pdf}."""
    fmt = (fmt or "csv").lower()
    if fmt not in _EXPORTERS:
        raise ExportUnavailable(f"Unknown export format '{fmt}'. Use: {', '.join(_EXPORTERS)}.")
    fn, content_type = _EXPORTERS[fmt]
    data = fn(envelope)
    if isinstance(data, str):
        data = data.encode("utf-8")
    return data, content_type, f"dis_answer.{fmt}"
